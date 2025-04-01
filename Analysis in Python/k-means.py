import pandas as pd
import numpy as np
from sklearn.cluster import KMeans
from sklearn.preprocessing import StandardScaler
from sklearn.metrics import silhouette_score
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill
from scipy.spatial.distance import pdist, squareform
from kneed import KneeLocator


def gerar_relatorio_clusters(input_file, output_file='Relatorio_Clusters.xlsx', max_k=10):
    df_original = pd.read_excel(input_file)  # base original intacta
    df = df_original.copy()  # base de trabalho para o clustering

    # Apenas para análise dos clusters, sem colunas de ID
    df = df[[col for col in df.columns if 'ID' not in col.upper()]]
    numeric_df = df.select_dtypes(include=[np.number]).dropna()
    scaler = StandardScaler()
    scaled_data = scaler.fit_transform(numeric_df)

    inertia = []
    for k in range(1, max_k + 1):
        kmeans = KMeans(n_clusters=k, random_state=42)
        kmeans.fit(scaled_data)
        inertia.append(kmeans.inertia_)

    # Novo método do cotovelo com KneeLocator
    kneedle = KneeLocator(range(1, max_k + 1), inertia, curve='convex', direction='decreasing')
    k_opt = kneedle.elbow if kneedle.elbow else 2

    kmeans_final = KMeans(n_clusters=k_opt, random_state=42)
    clusters = kmeans_final.fit_predict(scaled_data)
    df['Cluster'] = -1
    df.loc[numeric_df.index, 'Cluster'] = clusters

    silhouette = silhouette_score(scaled_data, clusters)

    df_numeric_with_cluster = numeric_df.copy()
    df_numeric_with_cluster['Cluster'] = clusters
    resumo_clusters = df_numeric_with_cluster.groupby('Cluster').agg(['mean', 'std', 'max', 'min'])

    for col in numeric_df.columns:
        resumo_clusters[(col, 'range')] = resumo_clusters[(col, 'max')] - resumo_clusters[(col, 'min')]

    cv_clusters = resumo_clusters.copy()
    for col in numeric_df.columns:
        cv_clusters[(col, 'cv')] = resumo_clusters[(col, 'std')] / resumo_clusters[(col, 'mean')]

    media_geral = numeric_df.mean()
    cluster_sizes = df['Cluster'].value_counts().sort_index()
    cluster_props = (cluster_sizes / len(df) * 100).round(2)

    dispersoes = {}
    for c in range(k_opt):
        pontos_cluster = scaled_data[clusters == c]
        centroide = kmeans_final.cluster_centers_[c]
        dispersao = np.mean(np.linalg.norm(pontos_cluster - centroide, axis=1))
        dispersoes[c] = dispersao
    media_disp = np.mean(list(dispersoes.values()))
    std_disp = np.std(list(dispersoes.values()))

    dist_matrix = pd.DataFrame(
        squareform(pdist(kmeans_final.cluster_centers_)),
        columns=[f"C{i}" for i in range(k_opt)],
        index=[f"C{i}" for i in range(k_opt)]
    )

    # Métrica: Separação média entre clusters
    separacao_media = dist_matrix.values[np.triu_indices_from(dist_matrix.values, k=1)].mean()

    # Métrica: Densidade relativa
    densidades = {c: round(cluster_sizes[c] / dispersoes[c], 4) for c in dispersoes}

    # Métrica: Variação no tamanho dos clusters
    media_tam = np.mean(list(cluster_sizes))
    std_tam = np.std(list(cluster_sizes))

    destaques = {}
    for cluster in sorted(df['Cluster'].unique()):
        cluster_data = numeric_df[df['Cluster'] == cluster]
        media_cluster = cluster_data.mean()
        comparacao = media_cluster - media_geral
        destaques[cluster] = {
            'fortes': comparacao[comparacao > 0].sort_values(ascending=False),
            'fracos': comparacao[comparacao < 0].sort_values()
        }

    linhas = []
    now = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    linhas.append(["Relatório de Clusters", ""])
    linhas.append(["Data e hora da análise:", now])
    linhas.append(["Número de Clusters:", k_opt])
    linhas.append(["Total de Registros:", len(df)])
    linhas.append(["Índice de Silhueta:", round(silhouette, 4)])
    linhas.append(["Média da Dispersão:", round(media_disp, 4)])
    linhas.append(["Desvio da Dispersão:", round(std_disp, 4)])
    linhas.append(["Separação Média entre Clusters:", round(separacao_media, 4)])
    linhas.append(["Desvio no Tamanho dos Clusters:", round(std_tam, 2)])
    linhas.append(["Variáveis Utilizadas:", ", ".join(numeric_df.columns)])
    linhas.append([])

    linhas.append(["Método do Cotovelo"])
    linhas.append(["K", "Inércia"])
    for idx, val in enumerate(inertia, start=1):
        linhas.append([idx, round(val, 2)])
    linhas.append([])

    linhas.append(["Tamanho dos Clusters"])
    linhas.append(["Cluster", "Quantidade", "% do Total"])
    for idx in cluster_sizes.index:
        linhas.append([idx, cluster_sizes[idx], cluster_props[idx]])
    linhas.append([])

    linhas.append(["Densidade Relativa por Cluster"])
    linhas.append(["Cluster", "Densidade"])
    for idx in densidades:
        linhas.append([idx, densidades[idx]])
    linhas.append([])

    linhas.append(["Dispersão Intra-Cluster"])
    linhas.append(["Cluster", "Dispersão Média"])
    for idx, val in dispersoes.items():
        linhas.append([idx, round(val, 4)])
    linhas.append([])

    linhas.append(["Distância entre Centros dos Clusters"])
    linhas.append([""] + list(dist_matrix.columns))
    for idx, row in dist_matrix.iterrows():
        linhas.append([idx] + [round(val, 4) for val in row.values])
    linhas.append([])

    linhas.append(["Médias, Desvios, CV e Amplitude por Cluster"])
    header = ["Variável"]
    for cluster in sorted(df['Cluster'].unique()):
        header.extend([f"Cluster {cluster} - Média", f"Desvio", f"CV", f"Amplitude"])
    linhas.append(header)
    for col in numeric_df.columns:
        row = [col]
        for cluster in sorted(df['Cluster'].unique()):
            media = resumo_clusters.loc[cluster, (col, 'mean')]
            desvio = resumo_clusters.loc[cluster, (col, 'std')]
            cv = cv_clusters.loc[cluster, (col, 'cv')]
            amplitude = resumo_clusters.loc[cluster, (col, 'range')]
            row.extend([
                round(media, 2),
                round(desvio, 2),
                round(cv, 2) if not np.isnan(cv) else '',
                round(amplitude, 2)
            ])
        linhas.append(row)
    linhas.append([])

    for cluster, info in destaques.items():
        linhas.append([f"Cluster {cluster} - Pontos Fortes"])
        linhas.append(["Variável", "Acima da Média"])
        for var, val in info['fortes'].items():
            linhas.append([var, round(val, 2)])
        linhas.append([])
        linhas.append([f"Cluster {cluster} - Pontos Fracos"])
        linhas.append(["Variável", "Abaixo da Média"])
        for var, val in info['fracos'].items():
            linhas.append([var, round(val, 2)])
        linhas.append([])

    relatorio_df = pd.DataFrame(linhas)
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        relatorio_df.to_excel(writer, index=False, header=False, sheet_name='Relatório de Clusters')

    wb = openpyxl.load_workbook(output_file)
    ws = wb.active

    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, (int, float)):
                if row[0].value == "K":
                    if cell.column == 2 and row[0].row > 0 and cell.value == round(inertia[k_opt - 1], 2):
                        cell.fill = green_fill
                elif row[0].value == "Dispersão Intra-Cluster":
                    continue
                elif row[0].value in range(k_opt):
                    if row[1].value == round(min(dispersoes.values()), 4):
                        row[1].fill = green_fill
                    elif row[1].value == round(max(dispersoes.values()), 4):
                        row[1].fill = red_fill
                elif isinstance(row[0].value, str) and row[0].value.startswith('C'):
                    valid_vals = [c for c in row[1:] if isinstance(c.value, (int, float)) and c.value > 0]
                    if valid_vals:
                        min_val = min(c.value for c in valid_vals)
                        max_val = max(c.value for c in valid_vals)
                        for c in valid_vals:
                            if c.value == min_val:
                                c.fill = green_fill
                            elif c.value == max_val:
                                c.fill = red_fill
                elif "CV" in str(cell.coordinate):
                    row_vals = [c.value for c in row if isinstance(c.value, (int, float))]
                    if cell.value == min(row_vals):
                        cell.fill = green_fill
                    elif cell.value == max(row_vals):
                        cell.fill = red_fill
                elif "Amplitude" in str(cell.coordinate):
                    row_vals = [c.value for c in row if isinstance(c.value, (int, float))]
                    if cell.value == max(row_vals):
                        cell.fill = red_fill
                    elif cell.value == min(row_vals):
                        cell.fill = green_fill
                elif row[0].value == "Separação Média entre Clusters:":
                    cell.fill = green_fill  # Quanto maior, melhor
                elif row[0].value == "Desvio no Tamanho dos Clusters:":
                    cell.fill = red_fill  # Quanto menor, melhor balanceamento
                elif row[0].value == "Densidade Relativa por Cluster":
                    if row[1].value == max(densidades.values()):
                        row[1].fill = green_fill
                    elif row[1].value == min(densidades.values()):
                        row[1].fill = red_fill
                elif isinstance(row[0].value, str) and "Pontos Fortes" in row[0].value:
                    if cell.value > 0:
                        cell.fill = green_fill
                elif isinstance(row[0].value, str) and "Pontos Fracos" in row[0].value:
                    if cell.value < 0:
                        cell.fill = red_fill

    wb.save(output_file)
    
    # Salvar a base original completa com a coluna de Cluster
    base_completa_com_clusters = df_original.copy()
    base_completa_com_clusters['Cluster'] = np.nan
    base_completa_com_clusters.loc[numeric_df.index, 'Cluster'] = clusters

    base_com_clusters_file = output_file.replace('.xlsx', '_com_base.xlsx')
    base_completa_com_clusters.to_excel(base_com_clusters_file, index=False)
    print(f"Base com clusters salva: {base_com_clusters_file}")

    print(f"Relatório de clusters gerado: {output_file}")


if __name__ == '__main__':
    gerar_relatorio_clusters('Database.xlsx')

