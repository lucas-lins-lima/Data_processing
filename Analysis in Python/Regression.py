import os
import pandas as pd
import numpy as np
import statsmodels.api as sm
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
from datetime import datetime

# ---------------------- EXPORTAÇÃO FORMATADA ----------------------
def exportar_resultado_formatado(titulo, df_model, reg_result, reg_result_model, tipo_regressao, var_dependente, variaveis_explicativas, nome_arquivo):

    pasta = "output/linear" if tipo_regressao == "Linear" else "output/logistica"
    os.makedirs(pasta, exist_ok=True)
    caminho = os.path.join(pasta, nome_arquivo)

    wb = Workbook()
    ws = wb.active
    ws.title = "Resultados"
    bold = Font(bold=True)
    inicio = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

    # Info da execução
    ws.append(["Relatório de Regressão"])
    ws.append(["Data e hora de criação:", inicio])
    ws.append(["Tipo de Regressão:", tipo_regressao])
    ws.append(["Variável dependente:", var_dependente])
    ws.append(["Variáveis explicativas:", ", ".join(variaveis_explicativas)])
    ws.append([])

    # Estatísticas descritivas
    descr = df_model.describe().T[['count', 'mean', 'std', 'min', 'max']]
    descr.reset_index(inplace=True)
    descr.columns = ['Variável', 'N', 'Média', 'Desvio Padrão', 'Mínimo', 'Máximo']
    ws.append(["Descriptive Statistics"])
    for r in dataframe_to_rows(descr, index=False, header=True):
        ws.append(r)

    ws.append([])

    # ---------------------- Metodologia com métricas ----------------------
    ws.append(["Metodologia"])

    if tipo_regressao == "Linear":
        ws.append(["Análise realizada via regressão linear. Intervalo de confiança de 95%."])
        ws.append([f"R²: {reg_result_model.rsquared:.3f} – Proporção da variância explicada pelas variáveis independentes."])
        ws.append([f"R² ajustado: {reg_result_model.rsquared_adj:.3f} – R² corrigido pelo número de variáveis no modelo."])
        ws.append([f"Erro padrão da regressão: {np.sqrt(reg_result_model.scale):.3f} – Mede a dispersão dos resíduos."])
        ws.append([f"Estatística F: {reg_result_model.fvalue:.3f} (p-valor: {reg_result_model.f_pvalue:.3f}) – Testa a significância do modelo como um todo."])
    else:
        ws.append(["Análise realizada via regressão logística. Intervalo de confiança de 95%."])
        ws.append([f"Pseudo R² (McFadden): {reg_result_model.prsquared:.3f} – Indica a qualidade do ajuste do modelo (0 a 1)."])
        ws.append([f"Log-Likelihood: {reg_result_model.llf:.3f} – Valor da log-verossimilhança, quanto maior (menos negativo), melhor."])
        ws.append([f"AIC: {reg_result_model.aic:.3f} – Critério de informação de Akaike (quanto menor, melhor)."])
        ws.append([f"BIC: {reg_result_model.bic:.3f} – Critério bayesiano de informação (penaliza modelos complexos)."])

    ws.append([])

    # Resumo da Amostra
    ws.append(["Resumo da Amostra"])
    ws.append(["Total de casos utilizados:", len(df_model)])
    ws.append(["Casos perdidos/removidos:", df_model.isna().sum().sum()])
    ws.append([])

    # Escala
    ws.append(["Escala da Variável Dependente"])
    escala = "0 = Não | 1 = Sim" if tipo_regressao == 'Logística' else "1 a 5 (ordinal)"
    ws.append([escala])
    ws.append([])

    # Resultados da Regressão
    ws.append(["Regression Results"])
    for i, r in enumerate(dataframe_to_rows(reg_result, index=False, header=True)):
        ws.append(r)
        if i > 0 and len(r) > 3:
            p = r[3]
            if isinstance(p, (int, float)) and p < 0.05:
                for c in range(1, len(r)+1):
                    ws.cell(row=ws.max_row, column=c).font = bold

    wb.save(caminho)
    print(f"✅ Exportado: {caminho}")

# ---------------------- REGRESSÕES ----------------------
def rodar_modelo_base(df, var_dependente, variaveis_explicativas, categorias=None, tipo='linear'):
    colunas = [var_dependente] + variaveis_explicativas
    df_model = df[colunas].copy()

    if categorias:
        df_model = pd.get_dummies(df_model, columns=categorias, drop_first=True)

    df_model = df_model.loc[:, ~df_model.columns.duplicated()]
    for col in df_model.columns:
        try:
            if df_model[col].dtype == bool:
                df_model[col] = df_model[col].astype(int)
            else:
                df_model[col] = pd.to_numeric(df_model[col], errors='coerce')
        except Exception as e:
            print(f"⚠️ Erro ao converter a coluna '{col}': {e}")

    df_model = df_model.dropna()
    y = df_model[var_dependente]
    X = df_model.drop(columns=[var_dependente])
    X = sm.add_constant(X)

    if tipo == 'linear':
        modelo = sm.OLS(y, X).fit()
        reg_result = pd.DataFrame({
            'Variável': modelo.params.index,
            'Coeficiente': modelo.params.values,
            'Erro Padrão': modelo.bse,
            'p-valor': modelo.pvalues,
            'IC 2.5%': modelo.conf_int()[0],
            'IC 97.5%': modelo.conf_int()[1]
        })
    elif tipo == 'logistica':
        if not set(y.unique()).issubset({0, 1}):
            raise ValueError(f"A variável '{var_dependente}' precisa ser binária (0 e 1).")
        modelo = sm.Logit(y, X).fit(disp=False)
        coef = modelo.params
        stderr = modelo.bse
        pvals = modelo.pvalues
        conf = modelo.conf_int()
        odds_ratio = np.exp(coef)
        b_positiva = [b if b > 0 else 0 for b in coef]
        reg_result = pd.DataFrame({
            'Variável': coef.index,
            'B': coef.values,
            'Erro Padrão': stderr,
            'p-valor': pvals,
            'Odds Ratio (Exp(B))': odds_ratio,
            'IC 2.5%': conf[0],
            'IC 97.5%': conf[1],
            'B_positiva (SE(B>0;B;0))': b_positiva
        })
    else:
        raise ValueError("Tipo de regressão inválido. Use 'linear' ou 'logistica'.")

    return modelo, df_model, reg_result

def criar_topbox_variaveis(df, coluna_original):
    return pd.DataFrame({
        'Topbox': (df[coluna_original] == 5).astype(int),
        'Top2Box': (df[coluna_original] >= 4).astype(int),
        'Bottombox': (df[coluna_original] == 1).astype(int),
        'Bottom2Box': (df[coluna_original] <= 2).astype(int)
    })

def rodar_regressoes_inteligentes(df, var_dependente, variaveis_explicativas, categorias=None):
    valores_unicos = df[var_dependente].dropna().unique()
    valores_unicos.sort()

    if set(valores_unicos).issubset({0, 1}):
        modelo, df_model, reg_result = rodar_modelo_base(df, var_dependente, variaveis_explicativas, categorias, tipo='logistica')
        return exportar_resultado_formatado(
            titulo=var_dependente,
            df_model=df_model,
            reg_result=reg_result,
            tipo_regressao='Logística',
            reg_result_model=modelo,
            var_dependente=var_dependente,
            variaveis_explicativas=variaveis_explicativas,
            nome_arquivo=f"Resultados_Regressao_Logistica_{var_dependente}.xlsx"
        )

    elif set(valores_unicos).issubset({1, 2, 3, 4, 5}):
        modelo, df_model, reg_result = rodar_modelo_base(df, var_dependente, variaveis_explicativas, categorias, tipo='linear')
        exportar_resultado_formatado(
            titulo=var_dependente,
            df_model=df_model,
            reg_result=reg_result,
            tipo_regressao='Linear',
            reg_result_model=modelo,
            var_dependente=var_dependente,
            variaveis_explicativas=variaveis_explicativas,
            nome_arquivo=f"Resultados_Regressao_Linear_{var_dependente}.xlsx"
        )

        topbox_df = criar_topbox_variaveis(df, var_dependente)
        df_ext = pd.concat([df.copy(), topbox_df], axis=1)

        for col in ['Topbox', 'Top2Box', 'Bottombox', 'Bottom2Box']:
            try:
                modelo, df_m, result = rodar_modelo_base(df_ext, col, variaveis_explicativas, categorias, tipo='logistica')
                exportar_resultado_formatado(
                    titulo=col,
                    df_model=df_m,
                    reg_result=result,
                    tipo_regressao='Logística',
                    reg_result_model=modelo,
                    var_dependente=col,
                    variaveis_explicativas=variaveis_explicativas,
                    nome_arquivo=f"Resultados_Regressao_Logistica_{col}.xlsx"
                )
            except Exception as e:
                print(f"Erro em {col}: {e}")
    else:
        raise ValueError(f"A variável '{var_dependente}' tem valores fora do padrão esperado (0–1 ou 1–5).")

def rodar_regressoes_binarias_em_lote_com_rotulo(df, lista_dependentes, rotulos_dict, variaveis_explicativas, categorias=None):
    for var in lista_dependentes:
        try:
            modelo, df_model, reg_result = rodar_modelo_base(df, var, variaveis_explicativas, categorias, tipo='logistica')
            exportar_resultado_formatado(
                titulo=rotulos_dict.get(var, var),
                df_model=df_model,
                reg_result=reg_result,
                tipo_regressao='Logística',
                reg_result_model=modelo,
                var_dependente=var,
                variaveis_explicativas=variaveis_explicativas,
                nome_arquivo=f"Resultados_Regressao_Logistica_{var}.xlsx"
            )
        except Exception as e:
            print(f"⚠️ Erro em '{var}': {e}")

def rodar_regressoes_lineares_em_lote_com_rotulo(df, lista_dependentes, rotulos_dict, variaveis_explicativas, categorias=None):
    for var in lista_dependentes:
        try:
            modelo, df_model, reg_result = rodar_modelo_base(df, var, variaveis_explicativas, categorias, tipo='linear')
            exportar_resultado_formatado(
                titulo=rotulos_dict.get(var, var),
                df_model=df_model,
                reg_result=reg_result,
                tipo_regressao='Linear',
                reg_result_model=modelo,
                var_dependente=var,
                variaveis_explicativas=variaveis_explicativas,
                nome_arquivo=f"Resultados_Regressao_Linear_{var}.xlsx"
            )
        except Exception as e:
            print(f"⚠️ Erro em '{var}': {e}")

# ---------------------- EXECUÇÃO ----------------------
df = pd.read_excel("Database.xlsx")

# Exemplo: regressão única
rodar_regressoes_inteligentes(
    df,
    var_dependente='Gosta_imoveis_Vitacon',
    variaveis_explicativas=[
        'FILTRO_Idade', 'Gênero', 'Região',
        'Usou_Vitacon', 'Conhecimento_mercado',
        'Prob_recomendar', 'Qualidade_percebida',
        'Custo_beneficio', 'Finalidade_imovel'
    ],
    categorias=['Gênero', 'Região', 'Finalidade_imovel']
)

# Exemplo: múltiplas binárias
rotulos_adjetivos = {
    "Frases_atributos_1": "Morar bem é uma prioridade pra mim",
    "Frases_atributos_2": "Meu imóvel precisa estar conectado ao meu ritmo de vida moderno",
    "Frases_atributos_3": "Busco um espaço que reflita minha personalidade e estilo",
    "Frases_atributos_4": "Vejo imóveis como uma forma de investimento para o futuro",
    "Frases_atributos_5": "Vejo a Vitacon como uma marca moderna e inovadora",
    "Frases_atributos_6": "Sou alguém que valoriza mobilidade urbana e soluções compactas",
    "Frases_atributos_7": "Valorizo imóveis com soluções que otimizem o meu dinheiro",
    "Frases_atributos_8": "Acho os imóveis da Vitacon com design inteligente e funcional",
    "Frases_atributos_9": "Prefiro pagar por menos espaço se for bem localizado",
    "Frases_atributos_10": "Gosto de marcas que pensam em sustentabilidade e tecnologia",
    "Frases_atributos_11": "Acredito em morar com praticidade, sem abrir mão do conforto",
    "Frases_atributos_12": "Me parece uma marca voltada para quem vive em grandes centros",
    "Frases_atributos_13": "Acredito em soluções compartilhadas e mais conscientes"
}
colunas_binarias = list(rotulos_adjetivos.keys())

rodar_regressoes_binarias_em_lote_com_rotulo(
    df=df,
    lista_dependentes=colunas_binarias,
    rotulos_dict=rotulos_adjetivos,
    variaveis_explicativas=[
        'FILTRO_Idade', 'Gênero', 'Região',
        'Conhece_Vitacon', 'Prob_recomendar', 'Custo_beneficio'
    ],
    categorias=['Gênero', 'Região']
)

# Exemplo: múltiplas lineares (se quiser rodar)
# rodar_regressoes_lineares_em_lote_com_rotulo(
#     df, lista_dependentes=['Satisfacao_1', 'Satisfacao_2'],
#     rotulos_dict={'Satisfacao_1': 'Serviço A', 'Satisfacao_2': 'Serviço B'},
#     variaveis_explicativas=[...], categorias=[...]
# )