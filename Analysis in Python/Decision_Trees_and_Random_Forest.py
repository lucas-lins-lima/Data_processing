import pandas as pd
import numpy as np
from sklearn.model_selection import train_test_split
from sklearn.ensemble import RandomForestClassifier
from sklearn.tree import DecisionTreeClassifier
from sklearn.preprocessing import LabelEncoder
from sklearn.metrics import classification_report, confusion_matrix, roc_auc_score
from sklearn.dummy import DummyClassifier
from scipy.stats import entropy
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill
from datetime import datetime

# carregar base
data = pd.read_excel("Database.xlsx")

# variável-alvo
alvo = "Gosta_imoveis_Vitacon"

# colunas que não serão usadas
colunas_ignorar = [
    "ID_respondente", "FILTRO_Idade", "FILTRO_Gênero", "FILTRO_Usou_Vitacon",
    "Frases_atributos_escolhidas", "Marcas_conhecidas", "Marcas_frequentes",
    "Marcas_mais_vendem", "Adjetivos_Vitacon", "Canais_informacao",
    "Criterios_escolha_imovel"
]

# selecionar variáveis explicativas
variaveis = [col for col in data.columns if col not in colunas_ignorar + [alvo]]

# remover linhas com alvo ausente
data = data.dropna(subset=[alvo])

# preencher valores faltantes
data[variaveis] = data[variaveis].fillna("desconhecido")

# codificar variáveis categóricas
for col in variaveis:
    if data[col].dtype == "object":
        data[col] = LabelEncoder().fit_transform(data[col])

X = data[variaveis]
y = data[alvo]

X_train, X_test, y_train, y_test = train_test_split(X, y, stratify=y, random_state=42)

rf = RandomForestClassifier(random_state=42, oob_score=True)
dt = DecisionTreeClassifier(random_state=42)
rf.fit(X_train, y_train)
dt.fit(X_train, y_train)

# Detalhamento
tipos = data[variaveis].dtypes.astype(str)
faltantes = (data[variaveis] == "desconhecido").sum()

importancias = pd.DataFrame({
    "Variável": X.columns,
    "Importância_RF (Redução de Impureza)": rf.feature_importances_,
    "Importância_DT (Baseada na Árvore)": dt.feature_importances_,
    "Tipo": [tipos[var] for var in X.columns],
    "Qtd_missing (antes preenchimento)": [faltantes[var] for var in X.columns]
})
importancias["Diferença_Importância"] = abs(importancias["Importância_RF (Redução de Impureza)"] - importancias["Importância_DT (Baseada na Árvore)"])
importancias["Rank_RF"] = importancias["Importância_RF (Redução de Impureza)"].rank(ascending=False).astype(int)
importancias = importancias.sort_values(by="Importância_RF (Redução de Impureza)", ascending=False)

relatorio_rf_dict = classification_report(y_test, rf.predict(X_test), output_dict=True, zero_division=0)
relatorio_dt_dict = classification_report(y_test, dt.predict(X_test), output_dict=True, zero_division=0)

wb = Workbook()
ws = wb.active
ws.title = "Análise Afinidade"

info_geral = [
    ["Relatório de Fatores que Influenciam a Afinidade com a Marca"],
    [f"Data e hora da análise: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"],
    [f"Variável-alvo: {alvo}"],
    [f"Número de registros utilizados: {len(data)}"],
    [f"Entropia da variável-alvo: {round(entropy(y.value_counts(normalize=True)), 4)}"],
    [""]
]
for row in info_geral:
    ws.append(row)

ws.append(["Distribuição da variável-alvo"])
for classe, qtd in y.value_counts().items():
    ws.append([f"Classe {classe}", qtd, f"{qtd/len(y):.2%}"])
ws.append([])

ws.append(["Importância das Variáveis (Random Forest e Decision Tree)"])
for r in dataframe_to_rows(importancias, index=False, header=True):
    ws.append(r)
ws.append([])

# BLOCO 3: Métricas adicionais
ws.append(["Métricas Adicionais dos Modelos"])
rf_depths = [est.get_depth() for est in rf.estimators_]
rf_leaves = [est.get_n_leaves() for est in rf.estimators_]
preds_all = np.array([tree.predict(X_test) for tree in rf.estimators_])
variancia_preds = preds_all.std(axis=0).mean()
ws.append(["Profundidade Média das Árvores - Random Forest", round(np.mean(rf_depths), 2)])
ws.append(["Número Médio de Folhas - Random Forest", round(np.mean(rf_leaves), 2)])
ws.append(["Variabilidade das previsões entre árvores - Random Forest", round(variancia_preds, 4)])
ws.append(["Profundidade da Árvore - Árvore de Decisão", dt.get_depth()])
ws.append(["Número de Folhas - Árvore de Decisão", dt.get_n_leaves()])
ws.append([])

ws.append(["Parâmetros - Random Forest", str(rf.get_params())])
ws.append(["Parâmetros - Decision Tree", str(dt.get_params())])
ws.append([])

# baseline
baseline = DummyClassifier(strategy="most_frequent")
baseline.fit(X_train, y_train)
acc_base = baseline.score(X_test, y_test)
acc_rf = rf.score(X_test, y_test)
acc_dt = dt.score(X_test, y_test)
ws.append(["Acurácia - Modelo Base (maioria)", acc_base])
ws.append(["Ganho de acurácia - Random Forest", acc_rf - acc_base])
ws.append(["Ganho de acurácia - Decision Tree", acc_dt - acc_base])
ws.append([])

if len(y.unique()) == 2:
    auc_rf = roc_auc_score(y_test, rf.predict_proba(X_test)[:, 1])
    auc_dt = roc_auc_score(y_test, dt.predict_proba(X_test)[:, 1])
    gini_rf = 2 * auc_rf - 1
    gini_dt = 2 * auc_dt - 1
    ws.append(["AUC (ROC) - Random Forest", auc_rf])
    ws.append(["Gini - Random Forest", gini_rf])
    ws.append(["AUC (ROC) - Decision Tree", auc_dt])
    ws.append(["Gini - Decision Tree", gini_dt])
    ws.append([])

# Matrizes de confusão
def matriz_confusao_normalizada(titulo, y_true, y_pred):
    ws.append([titulo])
    cm = confusion_matrix(y_true, y_pred)
    cm_norm = confusion_matrix(y_true, y_pred, normalize='true')
    ws.append(["Real vs Predito"])
    ws.append([""] + [f"Pred_{i}" for i in range(cm.shape[1])])
    for i, row in enumerate(cm):
        ws.append([f"Real_{i}"] + row.tolist())
    ws.append([])
    ws.append(["Matriz Normalizada"])
    ws.append([""] + [f"Pred_{i}" for i in range(cm.shape[1])])
    for i, row in enumerate(cm_norm):
        ws.append([f"Real_{i}"] + [round(x, 2) for x in row.tolist()])
    ws.append([])

matriz_confusao_normalizada("Matriz de Confusão - Random Forest", y_test, rf.predict(X_test))
matriz_confusao_normalizada("Matriz de Confusão - Decision Tree", y_test, dt.predict(X_test))

# Coloração analítica refinada
fill_verde = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
fill_amarelo = PatternFill(start_color="FFF59D", end_color="FFF59D", fill_type="solid")
fill_vermelho = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

faixas_metricas = {
    "acurácia": (0.80, 0.65),
    "ganho": (0.10, 0.05),
    "auc": (0.75, 0.60),
    "gini": (0.50, 0.30),
    "f1-score": (0.75, 0.60),
    "precision": (0.75, 0.60),
    "recall": (0.75, 0.60),
    "profundidade": (10, 5),
    "folhas": (20, 10),
    "variabilidade": (0.10, 0.20)  # menor é melhor nesse caso
}

def aplicar_cor_personalizada():
    for row in ws.iter_rows():
        for i, cell in enumerate(row):
            if isinstance(cell.value, float) and i > 0:
                metrica_nome = str(row[i-1].value).lower()
                for chave in faixas_metricas:
                    if chave in metrica_nome:
                        top, mid = faixas_metricas[chave]
                        if "variabilidade" in chave:
                            if cell.value <= top:
                                cell.fill = fill_verde
                            elif cell.value <= mid:
                                cell.fill = fill_amarelo
                            else:
                                cell.fill = fill_vermelho
                        else:
                            if cell.value >= top:
                                cell.fill = fill_verde
                            elif cell.value >= mid:
                                cell.fill = fill_amarelo
                            else:
                                cell.fill = fill_vermelho

aplicar_cor_personalizada()

wb.save("Relatorio_Afinidade_Marca.xlsx")